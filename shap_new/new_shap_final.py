import os
import pandas as pd
import joblib
import numpy as np
import shap
from openpyxl import load_workbook  # 加载现有Excel工作簿用
from utils import create_empty_ls
import time

# 加载特征数据
file_path = 'train_set_new.xlsx'
df = pd.read_excel(file_path, index_col=0)
data_feature_names = ['MagpieData minimum Number', 'MagpieData maximum Number', 'MagpieData range Number',
                      'MagpieData mean Number', 'MagpieData avg_dev Number', 'MagpieData mode Number',
                      'MagpieData minimum MendeleevNumber', 'MagpieData maximum MendeleevNumber',
                      'MagpieData range MendeleevNumber', 'MagpieData mean MendeleevNumber',
                      'MagpieData avg_dev MendeleevNumber', 'MagpieData mode MendeleevNumber',
                      'MagpieData minimum AtomicWeight', 'MagpieData maximum AtomicWeight',
                      'MagpieData range AtomicWeight', 'MagpieData mean AtomicWeight',
                      'MagpieData avg_dev AtomicWeight', 'MagpieData mode AtomicWeight',  'MagpieData mean MeltingT',
                      'MagpieData avg_dev MeltingT', 'MagpieData mode MeltingT', 'MagpieData minimum Column',
                      'MagpieData maximum Column', 'MagpieData range Column', 'MagpieData mean Column',
                      'MagpieData avg_dev Column', 'MagpieData mode Column', 'MagpieData minimum Row',
                      'MagpieData maximum Row', 'MagpieData range Row', 'MagpieData mean Row', 'MagpieData avg_dev Row',
                      'MagpieData mode Row', 'MagpieData minimum CovalentRadius', 'MagpieData maximum CovalentRadius',
                      'MagpieData range CovalentRadius', 'MagpieData mean CovalentRadius',
                      'MagpieData avg_dev CovalentRadius', 'MagpieData mode CovalentRadius',
                      'MagpieData minimum Electronegativity', 'MagpieData maximum Electronegativity',
                      'MagpieData range Electronegativity', 'MagpieData mean Electronegativity',
                      'MagpieData avg_dev Electronegativity', 'MagpieData mode Electronegativity',
                      'MagpieData minimum NsValence', 'MagpieData maximum NsValence', 'MagpieData range NsValence',
                      'MagpieData mean NsValence', 'MagpieData avg_dev NsValence', 'MagpieData mode NsValence',
                      'MagpieData minimum NpValence', 'MagpieData maximum NpValence', 'MagpieData range NpValence',
                      'MagpieData mean NpValence', 'MagpieData avg_dev NpValence', 'MagpieData mode NpValence',
                      'MagpieData minimum NdValence', 'MagpieData maximum NdValence', 'MagpieData range NdValence',
                      'MagpieData mean NdValence', 'MagpieData avg_dev NdValence', 'MagpieData mode NdValence',
                      'MagpieData minimum NfValence', 'MagpieData maximum NfValence', 'MagpieData range NfValence',
                      'MagpieData mean NfValence', 'MagpieData avg_dev NfValence', 'MagpieData mode NfValence',
                      'MagpieData minimum NValence', 'MagpieData maximum NValence', 'MagpieData range NValence',
                      'MagpieData mean NValence', 'MagpieData avg_dev NValence', 'MagpieData mode NValence',
                      'MagpieData minimum NsUnfilled', 'MagpieData maximum NsUnfilled', 'MagpieData range NsUnfilled',
                      'MagpieData mean NsUnfilled', 'MagpieData avg_dev NsUnfilled', 'MagpieData mode NsUnfilled',
                      'MagpieData minimum NpUnfilled', 'MagpieData maximum NpUnfilled', 'MagpieData range NpUnfilled',
                      'MagpieData mean NpUnfilled', 'MagpieData avg_dev NpUnfilled', 'MagpieData mode NpUnfilled',
                      'MagpieData minimum NdUnfilled', 'MagpieData maximum NdUnfilled', 'MagpieData range NdUnfilled',
                      'MagpieData mean NdUnfilled', 'MagpieData avg_dev NdUnfilled', 'MagpieData mode NdUnfilled',
                      'MagpieData minimum NfUnfilled', 'MagpieData maximum NfUnfilled', 'MagpieData range NfUnfilled',
                      'MagpieData mean NfUnfilled', 'MagpieData avg_dev NfUnfilled', 'MagpieData mode NfUnfilled',
                      'MagpieData minimum NUnfilled', 'MagpieData maximum NUnfilled', 'MagpieData range NUnfilled',
                      'MagpieData mean NUnfilled', 'MagpieData avg_dev NUnfilled', 'MagpieData mode NUnfilled',
                      'MagpieData minimum GSvolume_pa', 'MagpieData maximum GSvolume_pa',
                      'MagpieData range GSvolume_pa', 'MagpieData mean GSvolume_pa', 'MagpieData avg_dev GSvolume_pa',
                      'MagpieData mode GSvolume_pa', 'MagpieData minimum GSbandgap', 'MagpieData maximum GSbandgap',
                      'MagpieData range GSbandgap', 'MagpieData mean GSbandgap', 'MagpieData avg_dev GSbandgap',
                      'MagpieData mode GSbandgap', 'MagpieData minimum GSmagmom', 'MagpieData maximum GSmagmom',
                      'MagpieData range GSmagmom', 'MagpieData mean GSmagmom', 'MagpieData avg_dev GSmagmom',
                      'MagpieData mode GSmagmom', 'MagpieData minimum SpaceGroupNumber',
                      'MagpieData maximum SpaceGroupNumber', 'MagpieData range SpaceGroupNumber',
                      'MagpieData mean SpaceGroupNumber', 'MagpieData avg_dev SpaceGroupNumber',
                      'MagpieData mode SpaceGroupNumber', 'Yang delta', 'Yang omega', 'APE mean',
                      'Radii local mismatch', 'Radii gamma', 'Configuration entropy', 'Atomic weight mean',
                      'Total weight', 'Lambda entropy', 'Electronegativity delta', 'Electronegativity local mismatch',
                      'VEC mean', 'Mixing enthalpy', 'Mean cohesive energy', 'Interant electrons',
                      'Interant s electrons', 'Interant p electrons', 'Interant d electrons', 'Interant f electrons',
                      'Shear modulus mean', 'Shear modulus delta', 'Shear modulus local mismatch',
                      'Shear modulus strength model', 'H fraction', 'He fraction', 'Li fraction', 'Be fraction',
                      'B fraction', 'C fraction', 'N fraction', 'O fraction', 'F fraction', 'Ne fraction',
                      'Na fraction', 'Mg fraction', 'Al fraction', 'Si fraction', 'P fraction', 'S fraction',
                      'Cl fraction', 'Ar fraction', 'K fraction', 'Ca fraction', 'Sc fraction', 'Ti fraction',
                      'V fraction', 'Cr fraction', 'Mn fraction', 'Fe fraction', 'Co fraction', 'Ni fraction',
                      'Cu fraction', 'Zn fraction', 'Ga fraction', 'Ge fraction', 'As fraction', 'Se fraction',
                      'Br fraction', 'Kr fraction', 'Rb fraction', 'Sr fraction', 'Y fraction', 'Zr fraction',
                      'Nb fraction', 'Mo fraction', 'Tc fraction', 'Ru fraction', 'Rh fraction', 'Pd fraction',
                      'Ag fraction', 'Cd fraction', 'In fraction', 'Sn fraction', 'Sb fraction', 'Te fraction',
                      'I fraction', 'Xe fraction', 'Cs fraction', 'Ba fraction', 'La fraction', 'Ce fraction',
                      'Pr fraction', 'Nd fraction', 'Pm fraction', 'Sm fraction', 'Eu fraction', 'Gd fraction',
                      'Tb fraction', 'Dy fraction', 'Ho fraction', 'Er fraction', 'Tm fraction', 'Yb fraction',
                      'Lu fraction', 'Hf fraction', 'Ta fraction', 'W fraction', 'Re fraction', 'Os fraction',
                      'Ir fraction', 'Pt fraction', 'Au fraction', 'Hg fraction', 'Tl fraction', 'Pb fraction',
                      'Bi fraction', 'Po fraction', 'At fraction', 'Rn fraction', 'Fr fraction', 'Ra fraction',
                      'Ac fraction', 'Th fraction', 'Pa fraction', 'U fraction', 'Np fraction', 'Pu fraction',
                      'Am fraction', 'Cm fraction', 'Bk fraction', 'Cf fraction', 'Es fraction', 'Fm fraction',
                      'Md fraction', 'No fraction', 'Lr fraction', 'mean AtomicWeight', 'mean Column', 'mean Row',
                      'range Number', 'mean Number', 'range AtomicRadius', 'mean AtomicRadius',
                      'range Electronegativity', 'mean Electronegativity', 'avg s valence electrons',
                      'avg p valence electrons', 'avg d valence electrons', 'avg f valence electrons',
                      'frac s valence electrons', 'frac p valence electrons', 'frac d valence electrons',
                      'frac f valence electrons', 'Grain Size','Habit Plane', 'Calculated Solid Solution',
                      'Calculated Grain Boundary', 'Calculated Yield Strength']  # 特征名称列表
# 模型路径和问题配置
problems = {
    'qf': {
        'index_excel': 'index/qf_index.xlsx',
        'models': ['rf', 'xgboost']
    },
    'kl': {
        'index_excel': 'index/kl_index.xlsx',
        'models': ['rf', 'xgboost']
    }
}

# 函数: 处理模型
def process_models(problem, model_type, data_features_part):
    path = f'models/{problem}/{model_type}'
    print(f'Processing SHAP results for {problem}/{model_type}')
    index_excel_path = problems[problem]['index_excel']

    for dirpath, dirnames, filenames in os.walk(path):
        model_num = len(filenames)
        fold_ls = create_empty_ls(model_num)
        k_fold_df = pd.read_excel(f'index/{problem}_index.xlsx', sheet_name='k-fold', engine='openpyxl')

        overall_shap_df = pd.DataFrame()
        tic = time.time()

        for i, file in enumerate(filenames):
            print(f'Processing SHAP results for model {i + 1}')
            fold_ls[i] = list(k_fold_df.iloc[i][~k_fold_df.iloc[i].isna()])
            fold_ls[i].pop(0)  # 移除第一个非数据列元素
            x_test = data_features_part.iloc[fold_ls[i], :]

            model_file = os.path.join(dirpath, file)
            model = joblib.load(model_file)

            explainer = shap.TreeExplainer(model)
            shap_values = explainer.shap_values(x_test)
            shap.initjs()

            # 测试数据与SHAP值DataFrame
            df0 = pd.DataFrame(np.array(x_test), columns=x_test.columns)
            df_shap_values = pd.DataFrame(np.array(shap_values), columns=x_test.columns)

            # 计算每个模型的特征重要性
            shap_summary = np.abs(shap_values).mean(axis=0)
            shap_importance_df = pd.DataFrame({
                'Feature Name': data_features_part.columns.tolist(),
                'SHAP Value': shap_summary
            }).sort_values(by='SHAP Value', ascending=False)

            # 将当前模型的特征重要性添加到综合DF中
            overall_shap_df = pd.concat([overall_shap_df, shap_importance_df.assign(Model=f'{model_type}_{i + 1}')])

            # 添加执行保存逻辑
            book = load_workbook(index_excel_path)
            with pd.ExcelWriter(index_excel_path, engine='openpyxl', mode='a',if_sheet_exists = 'replace') as writer:
                writer.book = book

                df0.to_excel(writer, sheet_name=f'{problem}_{model_type}_original_{i + 1}', index=False)
                df_shap_values.to_excel(writer, sheet_name=f'{problem}_{model_type}_shap_{i + 1}', index=False)
                shap_importance_df.to_excel(writer, sheet_name=f'{problem}_{model_type}_{i + 1}_shap_importance',
                                            index=False)

        # 保存综合的特征重要性到Excel的一个单独sheet
        book = load_workbook(index_excel_path)
        with pd.ExcelWriter(index_excel_path, engine='openpyxl', mode='a',if_sheet_exists = 'replace') as writer:
            writer.book = book

            overall_shap_df.to_excel(writer, sheet_name=f'{problem}_{model_type}_1-5_shap_importance', index=False)

        toc = time.time()
        print(f'Processing time for 5 models: {toc - tic}s')

    return overall_shap_df

# 主循环
qf_overall_shap = pd.DataFrame()
kl_overall_shap = pd.DataFrame()

for problem in problems:
    data_features_part = df[data_feature_names]
    for model_type in problems[problem]['models']:
        if problem == 'qf':
            qf_overall_shap = pd.concat([qf_overall_shap, process_models(problem, model_type, data_features_part)])
        elif problem == 'kl':
            kl_overall_shap = pd.concat([kl_overall_shap, process_models(problem, model_type, data_features_part)])

# 整合两个问题的结果
qf_overall_shap = qf_overall_shap.groupby('Feature Name')['SHAP Value'].mean().reset_index().sort_values(by='SHAP Value', ascending=False)
kl_overall_shap = kl_overall_shap.groupby('Feature Name')['SHAP Value'].mean().reset_index().sort_values(by='SHAP Value', ascending=False)

# 保存到新的 sheet 中，分别保存 qf 和 kl 的结果到不同的 sheet 中
final_shap_excel_path = 'index/final_shap_importance.xlsx'
with pd.ExcelWriter(final_shap_excel_path, engine='openpyxl',mode='a',if_sheet_exists = 'replace') as writer:
    qf_overall_shap.to_excel(writer, sheet_name='qf_shap_importance', index=False)
    kl_overall_shap.to_excel(writer, sheet_name='kl_shap_importance', index=False)


print("Final SHAP importance results saved successfully!")
