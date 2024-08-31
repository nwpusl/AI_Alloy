import os.path

import numpy as np
import pandas as pd
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import r2_score, mean_absolute_percentage_error,mean_absolute_error
from utils import create_empty_ls,create_dir,mycopyfile,mycopydir
import joblib
from openpyxl import load_workbook
from xgboost import XGBRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import GradientBoostingRegressor  # 导入GradientBoostingRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.neighbors import KNeighborsRegressor
save_model=True
del_0qf=True #是否删除0屈服强度数据
save_index=True
# Regressor=GradientBoostingRegressor()
Regressor=RandomForestRegressor
label='qf'  # 问题名称
model_name1 = 'rf' #  模型名称记录
index_dir='index'  # 索引名称

main_path=f'models/{label}'

# 这里记录着分折的结果,方便后续读取序列作shap分析

index_excel=f'{index_dir}/qf_index.xlsx'
if not os.path.exists(index_dir):
    print('创建index文件夹')
    os.mkdir(index_dir)
else:
    print('index文件夹已存在')
if not os.path.exists(index_excel):
    print('生成index文件')
    df=pd.DataFrame()
    df.to_excel(index_excel)
else:
    print('index文件已存在')

create_dir(main_path,is_mainpath=True)
n_splits=5
ls1=['fold'+str(i) for i in range(1,n_splits+1)]
fold_dict=dict(zip(ls1,[0 for i in range(0,len(ls1))]))
# 设置随机种子以确保结果的可重复性
np.random.seed(200)

params = {
    'bootstrap': True,
    'ccp_alpha': 0.0,
    'criterion': 'squared_error',
    'max_depth': None,
    'max_features': 1.0,
    'max_leaf_nodes': None,
    'max_samples': None,
    'min_impurity_decrease': 0.0,
    'min_samples_leaf': 1,
    'min_samples_split': 2,
    'min_weight_fraction_leaf': 0.0,
    'n_estimators': 100,
    'n_jobs': None,
    'oob_score': False,
    'random_state': 200,
    'verbose': 0,
    'warm_start': False
}


# 读取 Excel 文件
df = pd.read_excel('Full.xlsx', index_col=0)  # 引入这一列之后，原本的第一列就是一个索引，第0列会从有意义的列开始
# 删除包含空值的行
if del_0qf:
    df = df[df['Yield_Strength'] != 0].reset_index(drop=True)
feature_names = df.columns[:-4]  # 特征名为屈服强度（倒数第四列）之前的部分
X = df.iloc[:, :-4]  # 特征: 最后四列之前的所有列
y = df['Yield_Strength']  # 目标: 倒数第四列
# print('X.columns:\n',X.columns)
# print('特征的形状:\n',len(feature_names))
# print('输入数据的形状：\n',X.shape)

# 将目标变量分箱为10个类别有利于进一步分折（但是预测标签不会这个改变）
bins = np.linspace(y.min(), y.max(), 11)
y_binned = np.digitize(y, bins)

# 初始化StratifiedKFold
skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=200)

model= Regressor(**params)  # 如果这里固定初始化模型删除掉的话可以在下面循环呢内添加随机初始化或者指定个别参数

# 初始化列表以存储每个折叠的分数
r2_scores = []
mape_scores = []
mae_scores = []
model_scores = []  # 存储每个分折的评估指标

index = 0
num_each_fold_train = np.floor((np.int32(X.shape[0]) - 1) * (n_splits - 1) / n_splits)
num_each_fold_test = (np.int32(X.shape[0]) - 1) // n_splits
for train_index, test_index in skf.split(X, y_binned):
    X_train, X_test = X.iloc[train_index[:int(num_each_fold_train)]], X.iloc[test_index[:int(num_each_fold_test)]]
    y_train, y_test = y.iloc[train_index[:int(num_each_fold_train)]], y.iloc[test_index[:int(num_each_fold_test)]]
    index = index + 1
    fold_dict['fold' + str(index)] = train_index[:int(num_each_fold_train)].astype(
        int)  # 记录训练集索引    # print(f'fold_{index}',fold_dict)
    # 训练模型
    model.fit(X_train, y_train)
    joblib.dump(model,   f'{main_path}/{model_name1}/{model_name1}{index}.pkl')
    print('保存模型文件')

    # 在测试集上进行预测
    y_pred = model.predict(X_test)

    # 计算并存储分数
    r2 = r2_score(y_test, y_pred)
    mape = mean_absolute_percentage_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)
    r2_scores.append(r2)
    mape_scores.append(mape)
    mae_scores.append(mae)

    # 存储每个分折的评估指标到字典中
    model_scores.append({
        "Fold": index,
        "R^2": r2,
        "MAPE": mape,
        "MAE": mae
    })

ls = list(fold_dict.values())
# print('ls',ls)
df = pd.DataFrame(ls)

book = load_workbook(index_excel)
with pd.ExcelWriter(index_excel, engine='openpyxl') as writer:
    writer.book = book
    df.to_excel(writer, sheet_name='k-fold')

# 计算最终模型的性能
final_r2 = np.mean(r2_scores)
final_mape = np.mean(mape_scores)
final_mae = np.mean(mae_scores)

# 添加最终模型性能到DataFrame中
model_scores.append({
    "Fold": "Final",
    "R^2": final_r2,
    "MAPE": final_mape,
    "MAE": final_mae
})

# 创建DataFrame来存储每个分折的评估指标和最终模型性能
model_scores_df = pd.DataFrame(model_scores)

save_path = "individual_model_scores.xlsx"
if not os.path.exists(save_path):
    df = pd.DataFrame()
    df.to_excel(save_path, index=False)
# 将DataFrame写入Excel文件
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_to_existing_excel(scores_df, file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        # 计算新数据的起始行
        start_row = ws.max_row + 1
        # 将数据写入工作表
        for row in dataframe_to_rows(scores_df, index=False, header=True):
            ws.append(row)
        # 保存工作表
        wb.save(file_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建新文件并将数据写入
        scores_df.to_excel(file_path, index=False, sheet_name=sheet_name)

write_to_existing_excel(model_scores_df, save_path, 'qf_rf')