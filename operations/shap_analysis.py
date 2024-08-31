import pandas as pd
import numpy as np
import os
from joblib import load
import shap
from openpyxl import load_workbook
import joblib
from utils import create_empty_ls
# data = pd.read_excel('FULL.xlsx')
data = pd.read_excel('FULL.xlsx', index_col=0)
feature_names = data.columns[:-4]
print('feature_names', len(feature_names))
data_features_part = data[feature_names]
model_num = 5
name = 'rf'
main_path = f'feature_importance/{name}'
path = f'models/{name}'
excel_output1 = f'feature_importance/{name}/{name}_tree_importance.xlsx'  # 输出Excel的文件名
excel_output2 = f'feature_importance/{name}/{name}_shap_importance.xlsx'  # 输出Excel的文件名
excel = f'index/{name}_index.xlsx'
# 准备在循环外部创建ExcelWriter，以便在一个文档中保存所有sheets
with pd.ExcelWriter(excel_output1) as writer:
    i = 0  # 模型计数器，从1开始
    feature_importances_final = np.zeros(len(list(feature_names)))
    print('size of feature_importance_initial:', len(feature_importances_final))
    for dirpath, dirnames, filenames in os.walk(path):
        model_num = len(filenames)
        for file in filenames:
            model_file = os.path.join(dirpath, file)
            model = load(model_file)
            # new_data=np.array(data.iloc[1, 1:-4]).reshape(-1,1)
            # print('new_data:',new_data)
            # print('data.shape:',new_data.shape)
            # y=model.predict(new_data)
            # print('预测值：',y)
            # Tree importance 部分
            tree_importance_values = model.feature_importances_
            feature_importances_final += model.feature_importances_
            tree_importance_df = pd.DataFrame({
                'Feature Name': feature_names,
                'Importance Value': tree_importance_values
            }).sort_values(by='Importance Value', ascending=False).head(15).reset_index(drop=True)
            tree_importance_df.to_excel(writer, sheet_name=f'model{i + 1} tree importance', index=False)
            i += 1  # 更新模型计数器
    # 输出全部结果中的的15个重要特征
    feature_importances_final /= model_num
    importance_dict = dict(zip(feature_importances_final, feature_names))
    importance_tuplelist = [(abs(value), feature) for (value, feature) in importance_dict.items()]
    importance_sorted = sorted(importance_tuplelist, reverse=True)
    # 输出排名靠前列表的序号
    x_feature = [j for (i, j) in importance_sorted]
    y_importance_value = [i for (i, j) in importance_sorted]
    print(f'{name}_tree_feature_names(top15):')
    for i in x_feature[:15]:
        print(i)
    print(f'average_{name}_feature_values(top15):')
    for i in y_importance_value[:15]:
        print(i)
    columns_mapping = {col: data_features_part.columns.get_loc(col) for col in x_feature[:15]}
    print("Mapping of x_feature names to columns in data_features_part:")
    for feature, col_index in columns_mapping.items():
        print(f"Feature: {feature} is in column index: {col_index}")
with pd.ExcelWriter(excel_output2) as writer:
    i = 0
    feature_shap_importance_final = np.zeros(len(list(feature_names)))
    for dirpath, dirnames, filenames in os.walk(path):
        for file in filenames:
            print(f'计算第{i + 1}个模型的shap值结果')
            model_num = len(filenames)
            fold_ls = create_empty_ls(model_num)
            k_fold_df = pd.read_excel(excel, sheet_name='k-fold', index_col=0)
            fold_ls[i] = list(np.array(k_fold_df.iloc[i]))
            # SHAP importance 部分
            x_train = data_features_part.iloc[fold_ls[i]]
            explainer = shap.TreeExplainer(model)
            model_file = os.path.join(dirpath, file)
            model = joblib.load(model_file)
            i += 1  # 更新模型计数器
            shap.initjs()
            shap_values = explainer(x_train)
            # print('shap_shape:',shap_values.shape)
            shap_values_array = shap_values.values
            # 保存shap值到excel表格
            # print('shap_shape:',shap_values_array.shape)
            df0 = pd.DataFrame(np.array(x_train))
            df = pd.DataFrame(np.array(shap_values_array))
            book = load_workbook(excel)
            # 输出特征重要性前十五项
            shap_summary = np.abs(shap_values_array).mean(axis=0)
            shap_importance_df = pd.DataFrame({
                'Feature Name': feature_names,
                'SHAP Value': shap_summary
            }).sort_values(by='SHAP Value', ascending=False).head(15).reset_index(drop=True)
            shap_importance_df.to_excel(writer, sheet_name=f'model{i} shap importance', index=False)
            feature_shap_importance_final += shap_summary
            # 写入到Excel
            with pd.ExcelWriter(excel) as writer1:
                writer1.book = book
                df0.to_excel(writer1, sheet_name=f'original{i}')
                df.to_excel(writer1, sheet_name=f'model{i}')
        shap_importance_dict = dict(zip(feature_shap_importance_final / model_num, feature_names))
        shap_importance_tuplelist = [(abs(value), feature) for (value, feature) in shap_importance_dict.items()]
        importance_sorted = sorted(shap_importance_tuplelist, reverse=True)
        # 输出排名靠前列表的序号
        x_feature = [j for (i, j) in importance_sorted]
        y_importance_value = [i for (i, j) in importance_sorted]
        print(f'{name}_shap_feature_names(top15):')
        for i in x_feature[:15]:
            print(i)
        print(f'average_{name}_feature_shap_values(top15):')
        for i in y_importance_value[:15]:
            print(i)
        # 找到原来特征所在的列号
        columns_mapping = {col: data_features_part.columns.get_loc(col) for col in x_feature[:15]}
        print("Mapping of x_feature names to columns in data_features_part:")
        for feature, col_index in columns_mapping.items():
            print(f"Feature: {feature} is in column index: {col_index}")
# 抗拉强度模型的SHAP分析
import pandas as pd
import numpy as np
import os
from joblib import load
import shap
from openpyxl import load_workbook
import joblib
data = pd.read_excel('FULL.xlsx', index_col=0)
data = data[data['Yield_Strength'] != 0].reset_index(drop=True)
print(data.columns)
feature_names = data.columns[0:-4]
print('feature_names:\n', feature_names)
print('feature_names_shape:\n', feature_names.shape)
data_features_part = data[feature_names]
print('feature_shape:\n', data_features_part.shape)
model_num = 5
name = 'xgboost'
main_path = f'feature_importance/{name}'
path = f'models/{name}'
excel_output1 = f'feature_importance/{name}/{name}_tree_importance.xlsx'  # 输出Excel的文件名
excel_output2 = f'feature_importance/{name}/{name}_shap_importance.xlsx'  # 输出Excel的文件名
excel = f'index/{name}_index.xlsx'
# 准备在循环外部创建ExcelWriter，以便在一个文档中保存所有sheets
with pd.ExcelWriter(excel_output1) as writer:
    i = 0  # 模型计数器，从1开始
    feature_importances_final = np.zeros(len(list(feature_names)))
    for dirpath, dirnames, filenames in os.walk(path):
        model_num = len(filenames)
        for file in filenames:
            model_file = os.path.join(dirpath, file)
            model = load(model_file)
            # Tree importance 部分
            tree_importance_values = model.feature_importances_
            feature_importances_final += model.feature_importances_
            tree_importance_df = pd.DataFrame({
                'Feature Name': feature_names,
                'Importance Value': tree_importance_values
            }).sort_values(by='Importance Value', ascending=False).head(15).reset_index(drop=True)
            tree_importance_df.to_excel(writer, sheet_name=f'model{i + 1} tree importance', index=False)
            i += 1  # 更新模型计数器
    # 输出全部结果中的的15个重要特征
    feature_importances_final /= model_num
    importance_dict = dict(zip(feature_importances_final, feature_names))
    importance_tuplelist = [(abs(value), feature) for (value, feature) in importance_dict.items()]
    importance_sorted = sorted(importance_tuplelist, reverse=True)
    # 输出排名靠前列表的序号
    x_feature = [j for (i, j) in importance_sorted]
    y_importance_value = [i for (i, j) in importance_sorted]
    print(f'{name}_tree_feature_names(top15):')
    for i in x_feature[:15]:
        print(i)
    print(f'average_{name}_feature_values(top15):')
    for i in y_importance_value[:15]:
        print(i)
    columns_mapping = {col: data_features_part.columns.get_loc(col) for col in x_feature[:15]}
    print("Mapping of x_feature names to columns in data_features_part:")
    for feature, col_index in columns_mapping.items():
        print(f"Feature: {feature} is in column index: {col_index}")
with pd.ExcelWriter(excel_output2) as writer:
    i = 0
    feature_shap_importance_final = np.zeros(len(list(feature_names)))
    for dirpath, dirnames, filenames in os.walk(path):
        for file in filenames:
            print(f'计算第{i + 1}个模型的shap值结果')
            model_num = len(filenames)
            fold_ls = create_empty_ls(model_num)
            k_fold_df = pd.read_excel(excel, sheet_name='k-fold').iloc[:, 1:]
            fold_ls[i] = list(np.array(k_fold_df.iloc[i]))
            # SHAP importance 部分
            x_train = data_features_part.iloc[fold_ls[i]]
            explainer = shap.TreeExplainer(model)
            model_file = os.path.join(dirpath, file)
            model = joblib.load(model_file)
            i += 1  # 更新模型计数器
            shap.initjs()
            shap_values = explainer(x_train)
            # print('shap_shape:',shap_values.shape)
            shap_values_array = shap_values.values
            # 保存shap值到excel表格
            # print('shap_shape:',shap_values_array.shape)
            df0 = pd.DataFrame(np.array(x_train))
            df = pd.DataFrame(np.array(shap_values_array))
            book = load_workbook(excel)
            # 输出特征重要性前十五项
            shap_summary = np.abs(shap_values_array).mean(axis=0)
            shap_importance_df = pd.DataFrame({
                'Feature Name': feature_names,
                'SHAP Value': shap_summary
            }).sort_values(by='SHAP Value', ascending=False).head(15).reset_index(drop=True)
            shap_importance_df.to_excel(writer, sheet_name=f'model{i} shap importance', index=False)
            feature_shap_importance_final += shap_summary
            # 写入到Excel
            with pd.ExcelWriter(excel) as writer1:
                writer1.book = book
                df0.to_excel(writer1, sheet_name=f'original{i}')
                df.to_excel(writer1, sheet_name=f'model{i}')
        shap_importance_dict = dict(zip(feature_shap_importance_final / model_num, feature_names))
        shap_importance_tuplelist = [(abs(value), feature) for (value, feature) in shap_importance_dict.items()]
        importance_sorted = sorted(shap_importance_tuplelist, reverse=True)
        # 输出排名靠前列表的序号
        x_feature = [j for (i, j) in importance_sorted]
        y_importance_value = [i for (i, j) in importance_sorted]
        print(f'{name}_shap_feature_names(top15):')
        for i in x_feature[:15]:
            print(i)
        print(f'average_{name}_feature_shap_values(top15):')
        for i in y_importance_value[:15]:
            print(i)
        # 找到原来特征所在的列号
        columns_mapping = {col: data_features_part.columns.get_loc(col) for col in x_feature[:15]}
        print("Mapping of x_feature names to columns in data_features_part:")
        for feature, col_index in columns_mapping.items():
            print(f"Feature: {feature} is in column index: {col_index}")

